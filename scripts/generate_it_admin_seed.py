#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import json
import sys
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from stream_territory_wikidata_sql import iter_rows, parse_sql_tuple


SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parent
DEFAULT_OUTPUT_PATH = SCRIPT_DIR / "data" / "it_admin_seed.tsv"
ISTAT_WORKBOOK_URL = "https://www.istat.it/storage/codici-unita-amministrative/Elenco-comuni-italiani.xlsx"
TERRITORY_SQL = ROOT / "liquibase/changelog/18-load-territory-from-wikidata.sql"
WIKIDATA_SPARQL_URL = "https://query.wikidata.org/sparql"
WIKIDATA_USER_AGENT = "geo-api-it-admin-seed/1.0 (local maintenance script)"
EXPECTED_COUNTS = {
    "region": 20,
    "province_or_equivalent": 110,
    "municipality": 7894,
}
REGION_QIDS_BY_CODE = {
    "01": "Q1216",   # Piedmont
    "02": "Q1222",   # Aosta Valley
    "03": "Q1210",   # Lombardy
    "04": "Q1237",   # Trentino-South Tyrol
    "05": "Q1243",   # Veneto
    "06": "Q1250",   # Friuli Venezia Giulia
    "07": "Q1256",   # Liguria
    "08": "Q1263",   # Emilia-Romagna
    "09": "Q1273",   # Tuscany
    "10": "Q1280",   # Umbria
    "11": "Q1279",   # Marche
    "12": "Q1282",   # Lazio
    "13": "Q1284",   # Abruzzo
    "14": "Q1443",   # Molise
    "15": "Q1438",   # Campania
    "16": "Q1447",   # Apulia / Puglia
    "17": "Q1452",   # Basilicata
    "18": "Q1458",   # Calabria
    "19": "Q1460",   # Sicily
    "20": "Q1462",   # Sardinia
}
UTS_SOURCE_BY_TYPE = {
    1: "seed.it_admin_province",
    2: "seed.it_admin_autonomous_province",
    3: "seed.it_admin_metropolitan_city",
    4: "seed.it_admin_free_municipal_consortium",
    5: "seed.it_admin_former_province_statistical_unit",
}
UTS_OVERRIDES_BY_CODE = {
    # ISTAT 2026 Sardinia uses freshly renumbered province-equivalent codes that
    # are not yet aligned with Wikidata P635.
    "113": ("Q106996869", "provincia della Gallura Nord-Est Sardegna"),
    "116": ("Q16223", "provincia dell'Ogliastra"),
    # The metropolitan city still exposes the legacy province P635 code 058.
    "258": ("Q18288160", "citta metropolitana di Roma Capitale"),
}
MUNICIPALITY_OVERRIDES_BY_CODE = {
    "115079": ("Q389318", "Ula Tirso"),
    "118066": ("Q285572", "Villa San Pietro"),
}


@dataclass(frozen=True)
class WikidataMatch:
    qid: str
    label: str
    ended: str | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Generate the curated Italy administrative seed used by sync_admin_territory.sh "
            "from the official ISTAT Elenco dei comuni italiani workbook, Wikidata P635 "
            "codes and the versioned territory snapshot."
        )
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"Output TSV path (default: {DEFAULT_OUTPUT_PATH})",
    )
    parser.add_argument(
        "--istat-workbook",
        default=ISTAT_WORKBOOK_URL,
        help="Official ISTAT workbook URL or local path.",
    )
    parser.add_argument(
        "--territory-sql",
        type=Path,
        default=TERRITORY_SQL,
        help=f"Path to 18-load-territory-from-wikidata.sql (default: {TERRITORY_SQL})",
    )
    return parser.parse_args()


def fetch_bytes(source: str) -> bytes:
    path = Path(source)
    if path.is_file():
        return path.read_bytes()

    request = urllib.request.Request(source, headers={"User-Agent": WIKIDATA_USER_AGENT})
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


def load_istat_rows(source: str) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    workbook_bytes = fetch_bytes(source)
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    municipalities: list[dict[str, str]] = []
    regions_by_code: dict[str, dict[str, str]] = {}
    uts_by_code: dict[str, dict[str, str | int]] = {}

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_index == 1:
            continue

        region_code = str(row[0]).zfill(2)
        uts_code = str(row[1]).zfill(3)
        municipality_code = str(row[4]).zfill(6)
        municipality_display_name = row[5]
        region_name = row[10]
        uts_name = row[11]
        uts_type = int(row[12])

        if not municipality_display_name or not region_name or not uts_name:
            raise RuntimeError(f"Missing mandatory ISTAT fields in row {row_index}: {row!r}")

        regions_by_code[region_code] = {
            "admin_code": region_code,
            "display_name": region_name,
        }

        existing_uts = uts_by_code.get(uts_code)
        uts_payload = {
            "admin_code": uts_code,
            "display_name": uts_name,
            "uts_type": uts_type,
            "parent_admin_code": region_code,
        }
        if existing_uts is not None and existing_uts != uts_payload:
            raise RuntimeError(f"Conflicting ISTAT UTS metadata for code {uts_code}: {existing_uts!r} != {uts_payload!r}")
        uts_by_code[uts_code] = uts_payload

        municipalities.append(
            {
                "admin_code": municipality_code,
                "display_name": municipality_display_name,
                "parent_admin_code": uts_code,
            }
        )

    if len(municipalities) != EXPECTED_COUNTS["municipality"]:
        raise RuntimeError(
            f"Unexpected Italy municipality count: {len(municipalities)} != {EXPECTED_COUNTS['municipality']}"
        )
    if len(regions_by_code) != EXPECTED_COUNTS["region"]:
        raise RuntimeError(f"Unexpected Italy region count: {len(regions_by_code)} != {EXPECTED_COUNTS['region']}")
    if len(uts_by_code) != EXPECTED_COUNTS["province_or_equivalent"]:
        raise RuntimeError(
            f"Unexpected Italy province-or-equivalent count: {len(uts_by_code)} != "
            f"{EXPECTED_COUNTS['province_or_equivalent']}"
        )

    regions = [regions_by_code[code] for code in sorted(regions_by_code)]
    uts_units = [uts_by_code[code] for code in sorted(uts_by_code)]
    municipalities.sort(key=lambda row: int(row["admin_code"]))
    return municipalities, regions, uts_units


def load_italy_territory_snapshot(
    territory_sql: Path,
) -> tuple[dict[str, str], dict[str, list[str]]]:
    if not territory_sql.is_file():
        raise FileNotFoundError(f"Missing territory snapshot SQL file: {territory_sql}")

    territory_name_by_qid: dict[str, str] = {}
    municipality_qids_by_name: dict[str, list[str]] = defaultdict(list)

    for line_number, row in iter_rows(territory_sql):
        if ", 'IT'," not in row:
            continue
        parsed = parse_sql_tuple(row, line_number)
        wikidata_id, name, type_code, country_iso = parsed[:4]
        if not wikidata_id:
            continue
        territory_name_by_qid.setdefault(wikidata_id, name)
        if type_code == "municipality" and wikidata_id not in municipality_qids_by_name[name]:
            municipality_qids_by_name[name].append(wikidata_id)

    return territory_name_by_qid, municipality_qids_by_name


def build_wikidata_query() -> str:
    return """
SELECT ?code ?item ?itemLabel ?ended WHERE {
  ?item wdt:P635 ?code .
  OPTIONAL { ?item wdt:P576 ?ended }
  SERVICE wikibase:label { bd:serviceParam wikibase:language "it,en". }
}
""".strip()


def fetch_wikidata_bindings() -> list[dict[str, dict[str, str]]]:
    query_url = WIKIDATA_SPARQL_URL + "?" + urllib.parse.urlencode({"query": build_wikidata_query(), "format": "json"})
    request = urllib.request.Request(
        query_url,
        headers={
            "Accept": "application/sparql-results+json",
            "User-Agent": WIKIDATA_USER_AGENT,
        },
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        payload = json.load(response)
    return payload["results"]["bindings"]


def load_wikidata_matches(requested_codes: set[str]) -> dict[str, list[WikidataMatch]]:
    raw_matches: dict[str, dict[str, WikidataMatch]] = defaultdict(dict)
    for row in fetch_wikidata_bindings():
        code = row["code"]["value"]
        if code not in requested_codes:
            continue
        qid = row["item"]["value"].rsplit("/", 1)[-1]
        raw_matches[code][qid] = WikidataMatch(
            qid=qid,
            label=row["itemLabel"]["value"],
            ended=row.get("ended", {}).get("value"),
        )
    return {
        code: list(matches_by_qid.values())
        for code, matches_by_qid in raw_matches.items()
    }


def territory_name_for_qid(
    territory_name_by_qid: dict[str, str],
    *,
    qid: str,
    fallback: str,
) -> str:
    return territory_name_by_qid.get(qid, fallback)


def resolve_uts(
    uts: dict[str, str | int],
    wikidata_matches: dict[str, list[WikidataMatch]],
    territory_name_by_qid: dict[str, str],
) -> tuple[str, str]:
    override = UTS_OVERRIDES_BY_CODE.get(uts["admin_code"])
    if override is not None:
        qid, fallback_name = override
        return qid, territory_name_for_qid(territory_name_by_qid, qid=qid, fallback=fallback_name)

    candidates = wikidata_matches.get(uts["admin_code"], [])
    current_candidates = [candidate for candidate in candidates if candidate.ended is None]

    chosen: WikidataMatch | None = None
    if len(current_candidates) == 1:
        chosen = current_candidates[0]
    elif len(candidates) == 1:
        chosen = candidates[0]

    if chosen is None:
        raise RuntimeError(
            f"Expected a unique Wikidata match for Italian province-equivalent {uts['admin_code']}, "
            f"got {candidates!r}"
        )

    return chosen.qid, territory_name_for_qid(
        territory_name_by_qid,
        qid=chosen.qid,
        fallback=chosen.label,
    )


def resolve_municipality(
    municipality: dict[str, str],
    wikidata_matches: dict[str, list[WikidataMatch]],
    territory_name_by_qid: dict[str, str],
    municipality_qids_by_name: dict[str, list[str]],
) -> tuple[str, str]:
    override = MUNICIPALITY_OVERRIDES_BY_CODE.get(municipality["admin_code"])
    if override is not None:
        qid, fallback_name = override
        return qid, territory_name_for_qid(territory_name_by_qid, qid=qid, fallback=fallback_name)

    display_name = municipality["display_name"]
    candidates = wikidata_matches.get(municipality["admin_code"], [])
    current_candidates = [candidate for candidate in candidates if candidate.ended is None]
    snapshot_qids = municipality_qids_by_name.get(display_name, [])

    exact_current_label = [
        candidate for candidate in current_candidates if candidate.label.casefold() == display_name.casefold()
    ]
    exact_any_label = [candidate for candidate in candidates if candidate.label.casefold() == display_name.casefold()]
    snapshot_overlap = [candidate for candidate in candidates if candidate.qid in snapshot_qids]

    chosen_qid: str | None = None
    fallback_name: str = display_name
    if len(exact_current_label) == 1:
        chosen_qid = exact_current_label[0].qid
        fallback_name = exact_current_label[0].label
    elif len(exact_any_label) == 1:
        chosen_qid = exact_any_label[0].qid
        fallback_name = exact_any_label[0].label
    elif len(current_candidates) == 1:
        chosen_qid = current_candidates[0].qid
        fallback_name = current_candidates[0].label
    elif len(candidates) == 1:
        chosen_qid = candidates[0].qid
        fallback_name = candidates[0].label
    elif len(snapshot_overlap) == 1:
        chosen_qid = snapshot_overlap[0].qid
        fallback_name = snapshot_overlap[0].label
    elif len(snapshot_qids) == 1:
        chosen_qid = snapshot_qids[0]
    else:
        raise RuntimeError(
            f"Expected a unique municipality match for Italian ISTAT code {municipality['admin_code']} "
            f"({display_name!r}), got code candidates={candidates!r} snapshot candidates={snapshot_qids!r}"
        )

    return chosen_qid, territory_name_for_qid(
        territory_name_by_qid,
        qid=chosen_qid,
        fallback=fallback_name,
    )


def build_seed_rows(
    municipalities: list[dict[str, str]],
    regions: list[dict[str, str]],
    uts_units: list[dict[str, str | int]],
    wikidata_matches: dict[str, list[WikidataMatch]],
    territory_name_by_qid: dict[str, str],
    municipality_qids_by_name: dict[str, list[str]],
) -> list[dict[str, str]]:
    seed_rows: list[dict[str, str]] = []

    for region in regions:
        qid = REGION_QIDS_BY_CODE.get(region["admin_code"])
        if qid is None:
            raise RuntimeError(f"Missing Italy region QID mapping for code {region['admin_code']}")
        seed_rows.append(
            {
                "level_code": "region",
                "admin_code": region["admin_code"],
                "display_name": region["display_name"],
                "territory_name": territory_name_for_qid(
                    territory_name_by_qid,
                    qid=qid,
                    fallback=region["display_name"],
                ),
                "territory_wikidata_id": qid,
                "territory_type": "region",
                "parent_level_code": "",
                "parent_admin_code": "",
                "source": "seed.it_admin_region",
            }
        )

    for uts in uts_units:
        uts_type = int(uts["uts_type"])
        qid, territory_name = resolve_uts(uts, wikidata_matches, territory_name_by_qid)
        source = UTS_SOURCE_BY_TYPE.get(uts_type)
        if source is None:
            raise RuntimeError(f"Unsupported Italy UTS type {uts_type} for code {uts['admin_code']}")
        seed_rows.append(
            {
                "level_code": "province_or_equivalent",
                "admin_code": uts["admin_code"],
                "display_name": uts["display_name"],
                "territory_name": territory_name,
                "territory_wikidata_id": qid,
                "territory_type": "province",
                "parent_level_code": "region",
                "parent_admin_code": uts["parent_admin_code"],
                "source": source,
            }
        )

    for municipality in municipalities:
        qid, territory_name = resolve_municipality(
            municipality,
            wikidata_matches,
            territory_name_by_qid,
            municipality_qids_by_name,
        )
        seed_rows.append(
            {
                "level_code": "municipality",
                "admin_code": municipality["admin_code"],
                "display_name": municipality["display_name"],
                "territory_name": territory_name,
                "territory_wikidata_id": qid,
                "territory_type": "municipality",
                "parent_level_code": "province_or_equivalent",
                "parent_admin_code": municipality["parent_admin_code"],
                "source": "seed.it_admin_municipality",
            }
        )

    counts = Counter(seed_row["level_code"] for seed_row in seed_rows)
    if dict(counts) != EXPECTED_COUNTS:
        raise RuntimeError(f"Unexpected Italy seed counts: {dict(counts)!r} != {EXPECTED_COUNTS!r}")

    sort_order = {"region": 1, "province_or_equivalent": 2, "municipality": 3}
    return sorted(seed_rows, key=lambda row: (sort_order[row["level_code"]], row["admin_code"]))


def write_seed(rows: list[dict[str, str]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "level_code",
                "admin_code",
                "display_name",
                "territory_name",
                "territory_wikidata_id",
                "territory_type",
                "parent_level_code",
                "parent_admin_code",
                "source",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    args = parse_args()
    municipalities, regions, uts_units = load_istat_rows(args.istat_workbook)
    territory_name_by_qid, municipality_qids_by_name = load_italy_territory_snapshot(args.territory_sql)
    requested_codes = {
        *(municipality["admin_code"] for municipality in municipalities),
        *(uts["admin_code"] for uts in uts_units),
    }
    wikidata_matches = load_wikidata_matches(requested_codes)
    seed_rows = build_seed_rows(
        municipalities,
        regions,
        uts_units,
        wikidata_matches,
        territory_name_by_qid,
        municipality_qids_by_name,
    )
    write_seed(seed_rows, args.output)
    print(f"Wrote {len(seed_rows)} Italy administrative rows to {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
