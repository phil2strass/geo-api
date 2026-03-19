#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import sys
import time
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_PATH = SCRIPT_DIR / "data" / "nl_admin_seed.tsv"
MUNICIPALITIES_XLSX_URL = (
    "https://www.cbs.nl/-/media/cbs/onze-diensten/methoden/classificaties/overig/"
    "gemeenten-alfabetisch-2026.xlsx"
)
WIKIDATA_SPARQL_URL = "https://query.wikidata.org/sparql"
WIKIDATA_USER_AGENT = "geo-api-nl-admin-seed/1.0 (local maintenance script)"
EXPECTED_COUNTS = {
    "province": 12,
    "municipality": 345,
}
PROVINCE_QIDS_BY_CODE = {
    "PV20": "Q752",   # Groningen
    "PV21": "Q770",   # Friesland / Fryslaan
    "PV22": "Q772",   # Drenthe
    "PV23": "Q773",   # Overijssel
    "PV24": "Q707",   # Flevoland
    "PV25": "Q775",   # Gelderland
    "PV26": "Q776",   # Utrecht
    "PV27": "Q701",   # Noord-Holland
    "PV28": "Q694",   # Zuid-Holland
    "PV29": "Q705",   # Zeeland
    "PV30": "Q1101",  # Noord-Brabant
    "PV31": "Q1093",  # Limburg
}
SPECIAL_MUNICIPALITIES = [
    ("9001", "GM9001", "Bonaire"),
    ("9002", "GM9002", "Sint Eustatius"),
    ("9003", "GM9003", "Saba"),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Generate the curated Netherlands administrative seed used by "
            "sync_admin_territory.sh from the official CBS 2026 municipality workbook, "
            "the CBS note about Caribbean public bodies and Wikidata CBS municipality codes."
        )
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"Output TSV path (default: {DEFAULT_OUTPUT_PATH})",
    )
    parser.add_argument(
        "--municipalities-xlsx",
        default=MUNICIPALITIES_XLSX_URL,
        help="Official CBS municipalities workbook URL or local path.",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=80,
        help="Number of CBS municipality codes per Wikidata SPARQL request (default: 80).",
    )
    parser.add_argument(
        "--sleep-seconds",
        type=float,
        default=1.0,
        help="Delay between Wikidata requests (default: 1.0).",
    )
    return parser.parse_args()


def fetch_bytes(source: str) -> bytes:
    path = Path(source)
    if path.is_file():
        return path.read_bytes()

    request = urllib.request.Request(source, headers={"User-Agent": WIKIDATA_USER_AGENT})
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


def fetch_csv(url: str) -> list[dict[str, str]]:
    request = urllib.request.Request(
        url,
        headers={
            "Accept": "text/csv",
            "User-Agent": WIKIDATA_USER_AGENT,
        },
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        return list(csv.DictReader(io.StringIO(response.read().decode("utf-8"))))


def load_cbs_rows(source: str) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    workbook_bytes = fetch_bytes(source)
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    worksheet = workbook["Gemeenten_alfabetisch"]

    municipalities: list[dict[str, str]] = []
    provinces_by_code: dict[str, dict[str, str]] = {}
    for row in worksheet.iter_rows(values_only=True):
        if not (
            isinstance(row[0], str)
            and row[0].isdigit()
            and isinstance(row[1], str)
            and row[1].startswith("GM")
        ):
            continue

        municipality = {
            "raw_code": row[0],
            "admin_code": row[1],
            "display_name": row[2],
            "province_raw_code": row[3],
            "province_admin_code": row[4],
            "province_name": row[5],
        }
        municipalities.append(municipality)
        provinces_by_code[row[4]] = {
            "admin_code": row[4],
            "display_name": row[5],
        }

    if len(municipalities) != 342:
        raise RuntimeError(f"Unexpected CBS municipality count: {len(municipalities)} != 342")
    if len(provinces_by_code) != EXPECTED_COUNTS["province"]:
        raise RuntimeError(
            f"Unexpected CBS province count: {len(provinces_by_code)} != {EXPECTED_COUNTS['province']}"
        )

    provinces = [provinces_by_code[code] for code in sorted(provinces_by_code)]
    return municipalities, provinces


def build_wikidata_query(codes: list[str]) -> str:
    values = " ".join(f'"{code}"' for code in codes)
    return f"""
SELECT ?code ?item ?itemLabel WHERE {{
  VALUES ?code {{ {values} }}
  VALUES ?class {{ wd:Q2039348 wd:Q3237519 }}
  ?item wdt:P382 ?code ;
        wdt:P31 ?class .
  SERVICE wikibase:label {{ bd:serviceParam wikibase:language "nl,en". }}
}}
""".strip()


def load_wikidata_matches(
    codes: list[str],
    *,
    chunk_size: int,
    sleep_seconds: float,
) -> dict[str, list[tuple[str, str]]]:
    matches: dict[str, list[tuple[str, str]]] = defaultdict(list)
    for start in range(0, len(codes), chunk_size):
        chunk = codes[start : start + chunk_size]
        query_url = WIKIDATA_SPARQL_URL + "?" + urllib.parse.urlencode({"query": build_wikidata_query(chunk)})
        for row in fetch_csv(query_url):
            qid = row["item"].rsplit("/", 1)[-1]
            matches[row["code"]].append((qid, row["itemLabel"]))
        if sleep_seconds > 0:
            time.sleep(sleep_seconds)
    return matches


def resolve_unique(code: str, matches: dict[str, list[tuple[str, str]]]) -> tuple[str, str]:
    candidates = matches.get(code, [])
    if len(candidates) != 1:
        raise RuntimeError(f"Expected exactly one Wikidata match for Dutch CBS code {code}, got {candidates!r}")
    return candidates[0]


def build_seed_rows(
    municipalities: list[dict[str, str]],
    provinces: list[dict[str, str]],
    wikidata_matches: dict[str, list[tuple[str, str]]],
) -> list[dict[str, str]]:
    seed_rows: list[dict[str, str]] = []

    for province in provinces:
        province_qid = PROVINCE_QIDS_BY_CODE.get(province["admin_code"])
        if province_qid is None:
            raise RuntimeError(f"Missing province QID mapping for {province['admin_code']}")
        seed_rows.append(
            {
                "level_code": "province",
                "admin_code": province["admin_code"],
                "display_name": province["display_name"],
                "territory_name": province["display_name"],
                "territory_wikidata_id": province_qid,
                "territory_type": "province",
                "parent_level_code": "",
                "parent_admin_code": "",
                "source": "seed.nl_admin_province",
            }
        )

    for municipality in municipalities:
        qid, territory_name = resolve_unique(municipality["raw_code"], wikidata_matches)
        seed_rows.append(
            {
                "level_code": "municipality",
                "admin_code": municipality["admin_code"],
                "display_name": municipality["display_name"],
                "territory_name": territory_name,
                "territory_wikidata_id": qid,
                "territory_type": "municipality",
                "parent_level_code": "province",
                "parent_admin_code": municipality["province_admin_code"],
                "source": "seed.nl_admin_municipality",
            }
        )

    for raw_code, admin_code, name in SPECIAL_MUNICIPALITIES:
        qid, territory_name = resolve_unique(raw_code, wikidata_matches)
        seed_rows.append(
            {
                "level_code": "municipality",
                "admin_code": admin_code,
                "display_name": name,
                "territory_name": territory_name,
                "territory_wikidata_id": qid,
                "territory_type": "municipality",
                "parent_level_code": "",
                "parent_admin_code": "",
                "source": "seed.nl_admin_special_municipality",
            }
        )

    counts = Counter(seed_row["level_code"] for seed_row in seed_rows)
    if dict(counts) != EXPECTED_COUNTS:
        raise RuntimeError(f"Unexpected Netherlands seed counts: {dict(counts)!r} != {EXPECTED_COUNTS!r}")
    return seed_rows


def write_seed(rows: list[dict[str, str]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "level_code",
                "admin_code",
                "display_name",
                "territory_name",
                "territory_wikidata_id",
                "territory_type",
                "parent_level_code",
                "parent_admin_code",
                "source",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    args = parse_args()
    municipalities, provinces = load_cbs_rows(args.municipalities_xlsx)
    wikidata_matches = load_wikidata_matches(
        [municipality["raw_code"] for municipality in municipalities] + [row[0] for row in SPECIAL_MUNICIPALITIES],
        chunk_size=args.chunk_size,
        sleep_seconds=args.sleep_seconds,
    )
    seed_rows = build_seed_rows(municipalities, provinces, wikidata_matches)
    write_seed(seed_rows, args.output)
    print(f"Wrote {len(seed_rows)} Netherlands administrative rows to {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
