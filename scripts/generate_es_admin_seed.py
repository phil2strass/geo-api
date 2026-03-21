#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import sys
import time
import unicodedata
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from stream_territory_wikidata_sql import iter_rows, parse_sql_tuple


SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parent
DEFAULT_OUTPUT_PATH = SCRIPT_DIR / "data" / "es_admin_seed.tsv"
INE_WORKBOOK_URL = "https://www.ine.es/daco/daco42/codmun/diccionario26.xlsx"
TERRITORY_SQL = ROOT / "liquibase/changelog/18-load-territory-from-wikidata.sql"
WIKIDATA_SPARQL_URL = "https://query.wikidata.org/sparql"
WIKIDATA_USER_AGENT = "geo-api-es-admin-seed/1.0 (local maintenance script)"
EXPECTED_COUNTS = {
    "autonomous_community_or_city": 19,
    "province": 52,
    "municipality": 8132,
}
AUTONOMOUS_METADATA_BY_CODE = {
    "01": {
        "display_name": "Andalucía",
        "territory_wikidata_id": "Q5783",
        "territory_name": "Andalusia",
        "territory_type": "region",
    },
    "02": {
        "display_name": "Aragón",
        "territory_wikidata_id": "Q4040",
        "territory_name": "Aragon",
        "territory_type": "region",
    },
    "03": {
        "display_name": "Asturias",
        "territory_wikidata_id": "Q3934",
        "territory_name": "Asturias",
        "territory_type": "region",
    },
    "04": {
        "display_name": "Illes Balears",
        "territory_wikidata_id": "Q107356467",
        "territory_name": "Balearic Islands",
        "territory_type": "region",
    },
    "05": {
        "display_name": "Canarias",
        "territory_wikidata_id": "Q5813",
        "territory_name": "Canary Islands",
        "territory_type": "region",
    },
    "06": {
        "display_name": "Cantabria",
        "territory_wikidata_id": "Q3946",
        "territory_name": "Cantabria",
        "territory_type": "region",
    },
    "07": {
        "display_name": "Castilla y León",
        "territory_wikidata_id": "Q5739",
        "territory_name": "Castile and León",
        "territory_type": "region",
    },
    "08": {
        "display_name": "Castilla-La Mancha",
        "territory_wikidata_id": "Q5748",
        "territory_name": "Castile–La Mancha",
        "territory_type": "region",
    },
    "09": {
        "display_name": "Cataluña",
        "territory_wikidata_id": "Q5705",
        "territory_name": "Catalonia",
        "territory_type": "region",
    },
    "10": {
        "display_name": "Comunitat Valenciana",
        "territory_wikidata_id": "Q5720",
        "territory_name": "Valencian Community",
        "territory_type": "region",
    },
    "11": {
        "display_name": "Extremadura",
        "territory_wikidata_id": "Q5777",
        "territory_name": "Extremadura",
        "territory_type": "region",
    },
    "12": {
        "display_name": "Galicia",
        "territory_wikidata_id": "Q3908",
        "territory_name": "Galicia",
        "territory_type": "region",
    },
    "13": {
        "display_name": "Comunidad de Madrid",
        "territory_wikidata_id": "Q5756",
        "territory_name": "Community of Madrid",
        "territory_type": "region",
    },
    "14": {
        "display_name": "Región de Murcia",
        "territory_wikidata_id": "Q5772",
        "territory_name": "Region of Murcia",
        "territory_type": "region",
    },
    "15": {
        "display_name": "Navarra",
        "territory_wikidata_id": "Q4018",
        "territory_name": "Navarre",
        "territory_type": "region",
    },
    "16": {
        "display_name": "País Vasco",
        "territory_wikidata_id": "Q3995",
        "territory_name": "Basque Country",
        "territory_type": "region",
    },
    "17": {
        "display_name": "La Rioja",
        "territory_wikidata_id": "Q5727",
        "territory_name": "La Rioja",
        "territory_type": "region",
    },
    "18": {
        "display_name": "Ceuta",
        "territory_wikidata_id": "Q5823",
        "territory_name": "Ceuta",
        "territory_type": "region",
    },
    "19": {
        "display_name": "Melilla",
        "territory_wikidata_id": "Q5831",
        "territory_name": "Melilla",
        "territory_type": "municipality",
    },
}
PROVINCE_DISPLAY_NAME_BY_CODE = {
    "01": "Álava",
    "02": "Albacete",
    "03": "Alicante",
    "04": "Almería",
    "05": "Ávila",
    "06": "Badajoz",
    "07": "Illes Balears",
    "08": "Barcelona",
    "09": "Burgos",
    "10": "Cáceres",
    "11": "Cádiz",
    "12": "Castellón",
    "13": "Ciudad Real",
    "14": "Córdoba",
    "15": "A Coruña",
    "16": "Cuenca",
    "17": "Girona",
    "18": "Granada",
    "19": "Guadalajara",
    "20": "Gipuzkoa",
    "21": "Huelva",
    "22": "Huesca",
    "23": "Jaén",
    "24": "León",
    "25": "Lleida",
    "26": "La Rioja",
    "27": "Lugo",
    "28": "Madrid",
    "29": "Málaga",
    "30": "Murcia",
    "31": "Navarra",
    "32": "Ourense",
    "33": "Asturias",
    "34": "Palencia",
    "35": "Las Palmas",
    "36": "Pontevedra",
    "37": "Salamanca",
    "38": "Santa Cruz de Tenerife",
    "39": "Cantabria",
    "40": "Segovia",
    "41": "Sevilla",
    "42": "Soria",
    "43": "Tarragona",
    "44": "Teruel",
    "45": "Toledo",
    "46": "Valencia",
    "47": "Valladolid",
    "48": "Bizkaia",
    "49": "Zamora",
    "50": "Zaragoza",
    "51": "Ceuta",
    "52": "Melilla",
}
PROVINCE_QID_OVERRIDES = {
    "07": "Q107356469",
    "51": "Q5823",
    "52": "Q5831",
}
MUNICIPALITY_QID_OVERRIDES: dict[str, str] = {}


@dataclass(frozen=True)
class WikidataMatch:
    qid: str
    label: str
    statement_ended: str | None
    item_ended: str | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Generate the curated Spain administrative seed used by sync_admin_territory.sh "
            "from the official INE municipality dictionary workbook, Wikidata P772 codes and "
            "the versioned territory snapshot."
        )
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"Output TSV path (default: {DEFAULT_OUTPUT_PATH})",
    )
    parser.add_argument(
        "--ine-workbook",
        default=INE_WORKBOOK_URL,
        help="Official INE municipality dictionary XLSX URL or local path.",
    )
    parser.add_argument(
        "--territory-sql",
        type=Path,
        default=TERRITORY_SQL,
        help=f"Path to 18-load-territory-from-wikidata.sql (default: {TERRITORY_SQL})",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=150,
        help="Number of official INE codes per Wikidata request (default: 150).",
    )
    parser.add_argument(
        "--sleep-seconds",
        type=float,
        default=0.1,
        help="Delay between Wikidata requests (default: 0.1).",
    )
    return parser.parse_args()


def fetch_bytes(source: str) -> bytes:
    path = Path(source)
    if path.is_file():
        return path.read_bytes()

    request = urllib.request.Request(source, headers={"User-Agent": WIKIDATA_USER_AGENT})
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


def load_ine_rows(source: str) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    workbook = load_workbook(io.BytesIO(fetch_bytes(source)), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    municipalities: list[dict[str, str]] = []
    provinces_by_code: dict[str, dict[str, str]] = {}
    autonomous_codes_seen: set[str] = set()

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_index <= 2:
            continue

        autonomous_code = str(row[0]).zfill(2) if row[0] is not None else ""
        province_code = str(row[1]).zfill(2) if row[1] is not None else ""
        municipality_suffix = str(row[2]).zfill(3) if row[2] is not None else ""
        display_name = str(row[4]).strip() if row[4] is not None else ""

        if not autonomous_code or not province_code or not municipality_suffix or not display_name:
            raise RuntimeError(f"Missing mandatory INE fields in row {row_index}: {row!r}")

        autonomous_codes_seen.add(autonomous_code)

        province_payload = {
            "admin_code": province_code,
            "display_name": PROVINCE_DISPLAY_NAME_BY_CODE[province_code],
            "parent_admin_code": autonomous_code,
        }
        existing_province = provinces_by_code.get(province_code)
        if existing_province is not None and existing_province != province_payload:
            raise RuntimeError(
                f"Conflicting autonomous-community mapping for province {province_code}: "
                f"{existing_province!r} != {province_payload!r}"
            )
        provinces_by_code[province_code] = province_payload

        municipalities.append(
            {
                "admin_code": province_code + municipality_suffix,
                "display_name": display_name,
                "parent_admin_code": province_code,
            }
        )

    if len(autonomous_codes_seen) != EXPECTED_COUNTS["autonomous_community_or_city"]:
        raise RuntimeError(
            "Unexpected Spain autonomous-community/autonomous-city count: "
            f"{len(autonomous_codes_seen)} != {EXPECTED_COUNTS['autonomous_community_or_city']}"
        )
    if len(provinces_by_code) != EXPECTED_COUNTS["province"]:
        raise RuntimeError(
            f"Unexpected Spain province count: {len(provinces_by_code)} != {EXPECTED_COUNTS['province']}"
        )
    if len(municipalities) != EXPECTED_COUNTS["municipality"]:
        raise RuntimeError(
            f"Unexpected Spain municipality count: {len(municipalities)} != {EXPECTED_COUNTS['municipality']}"
        )

    provinces = [provinces_by_code[code] for code in sorted(provinces_by_code)]
    municipalities.sort(key=lambda row: int(row["admin_code"]))
    return municipalities, provinces


def load_snapshot_metadata(territory_sql: Path) -> tuple[set[str], dict[str, str]]:
    if not territory_sql.is_file():
        raise FileNotFoundError(f"Missing territory snapshot SQL file: {territory_sql}")

    qids: set[str] = set()
    names_by_qid: dict[str, str] = {}
    for line_number, row in iter_rows(territory_sql):
        if ", 'ES'," not in row:
            continue
        parsed = parse_sql_tuple(row, line_number)
        wikidata_id, name, _type_code, _country_iso = parsed[:4]
        if not wikidata_id:
            continue
        qids.add(wikidata_id)
        names_by_qid.setdefault(wikidata_id, name)
    return qids, names_by_qid


def build_wikidata_query(codes: list[str]) -> str:
    values = " ".join(f'"{code}"' for code in codes)
    return f"""
SELECT ?code ?item ?itemLabel ?statementEnded ?itemEnded WHERE {{
  VALUES ?code {{ {values} }}
  ?item p:P772 ?codeStatement .
  ?codeStatement ps:P772 ?code .
  OPTIONAL {{ ?codeStatement pq:P582 ?statementEnded }}
  OPTIONAL {{ ?item wdt:P576 ?itemEnded }}
  SERVICE wikibase:label {{ bd:serviceParam wikibase:language "es,ca,gl,eu,en". }}
}}
""".strip()


def fetch_csv(query: str) -> list[dict[str, str]]:
    request = urllib.request.Request(
        WIKIDATA_SPARQL_URL + "?" + urllib.parse.urlencode({"query": query}),
        headers={
            "Accept": "text/csv",
            "User-Agent": WIKIDATA_USER_AGENT,
        },
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        return list(csv.DictReader(io.StringIO(response.read().decode("utf-8"))))


def load_wikidata_matches(
    codes: list[str],
    *,
    chunk_size: int,
    sleep_seconds: float,
) -> dict[str, list[WikidataMatch]]:
    matches: dict[str, dict[str, WikidataMatch]] = defaultdict(dict)
    for start in range(0, len(codes), chunk_size):
        chunk = codes[start : start + chunk_size]
        for row in fetch_csv(build_wikidata_query(chunk)):
            qid = row["item"].rsplit("/", 1)[-1]
            matches[row["code"]][qid] = WikidataMatch(
                qid=qid,
                label=row.get("itemLabel", ""),
                statement_ended=row.get("statementEnded") or None,
                item_ended=row.get("itemEnded") or None,
            )
        if sleep_seconds > 0:
            time.sleep(sleep_seconds)
    return {code: list(matches_by_qid.values()) for code, matches_by_qid in matches.items()}


def normalize_name(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    stripped = "".join(char for char in normalized if not unicodedata.combining(char))
    simplified = []
    for char in stripped.lower():
        if char.isalnum():
            simplified.append(char)
        else:
            simplified.append(" ")
    return " ".join("".join(simplified).split())


def resolve_unique(
    code: str,
    *,
    expected_name: str,
    matches: dict[str, list[WikidataMatch]],
    snapshot_qids: set[str],
    snapshot_names_by_qid: dict[str, str],
    manual_qid: str | None = None,
) -> tuple[str, str]:
    if manual_qid is not None:
        return manual_qid, snapshot_names_by_qid.get(manual_qid, expected_name)

    current_candidates = [
        candidate
        for candidate in matches.get(code, [])
        if candidate.statement_ended is None and candidate.item_ended is None
    ]
    if not current_candidates:
        raise RuntimeError(f"Missing current Wikidata match for Spain INE code {code}")

    snapshot_candidates = [candidate for candidate in current_candidates if candidate.qid in snapshot_qids]
    if len(snapshot_candidates) == 1:
        candidate = snapshot_candidates[0]
        return candidate.qid, snapshot_names_by_qid.get(candidate.qid, expected_name)
    if snapshot_candidates:
        current_candidates = snapshot_candidates

    expected_key = normalize_name(expected_name)
    name_matches = [
        candidate
        for candidate in current_candidates
        if normalize_name(candidate.label) == expected_key
        or normalize_name(snapshot_names_by_qid.get(candidate.qid, "")) == expected_key
    ]
    if len(name_matches) == 1:
        candidate = name_matches[0]
        return candidate.qid, snapshot_names_by_qid.get(candidate.qid, expected_name)

    if len(current_candidates) == 1:
        candidate = current_candidates[0]
        raise RuntimeError(
            f"Spain INE code {code} / {expected_name!r} only matched {candidate.qid}, "
            "but that QID is missing from the versioned territory snapshot"
        )

    raise RuntimeError(
        f"Expected exactly one current Wikidata match for Spain INE code {code} / {expected_name!r}, "
        f"got {current_candidates!r}"
    )


def build_seed_rows(
    municipalities: list[dict[str, str]],
    provinces: list[dict[str, str]],
    province_matches: dict[str, list[WikidataMatch]],
    municipality_matches: dict[str, list[WikidataMatch]],
    snapshot_qids: set[str],
    snapshot_names_by_qid: dict[str, str],
) -> list[dict[str, str]]:
    seed_rows: list[dict[str, str]] = []

    for autonomous_code in sorted(AUTONOMOUS_METADATA_BY_CODE):
        metadata = AUTONOMOUS_METADATA_BY_CODE[autonomous_code]
        seed_rows.append(
            {
                "level_code": "autonomous_community_or_city",
                "admin_code": autonomous_code,
                "display_name": metadata["display_name"],
                "territory_name": metadata["territory_name"],
                "territory_wikidata_id": metadata["territory_wikidata_id"],
                "territory_type": metadata["territory_type"],
                "parent_level_code": "",
                "parent_admin_code": "",
                "source": "seed.es_admin_autonomous_community_or_city",
            }
        )

    for province in provinces:
        qid, territory_name = resolve_unique(
            province["admin_code"],
            expected_name=province["display_name"],
            matches=province_matches,
            snapshot_qids=snapshot_qids,
            snapshot_names_by_qid=snapshot_names_by_qid,
            manual_qid=PROVINCE_QID_OVERRIDES.get(province["admin_code"]),
        )
        seed_rows.append(
            {
                "level_code": "province",
                "admin_code": province["admin_code"],
                "display_name": province["display_name"],
                "territory_name": territory_name,
                "territory_wikidata_id": qid,
                "territory_type": "province",
                "parent_level_code": "autonomous_community_or_city",
                "parent_admin_code": province["parent_admin_code"],
                "source": "seed.es_admin_province",
            }
        )

    for municipality in municipalities:
        qid, territory_name = resolve_unique(
            municipality["admin_code"],
            expected_name=municipality["display_name"],
            matches=municipality_matches,
            snapshot_qids=snapshot_qids,
            snapshot_names_by_qid=snapshot_names_by_qid,
            manual_qid=MUNICIPALITY_QID_OVERRIDES.get(municipality["admin_code"]),
        )
        seed_rows.append(
            {
                "level_code": "municipality",
                "admin_code": municipality["admin_code"],
                "display_name": municipality["display_name"],
                "territory_name": territory_name,
                "territory_wikidata_id": qid,
                "territory_type": "municipality",
                "parent_level_code": "province",
                "parent_admin_code": municipality["parent_admin_code"],
                "source": "seed.es_admin_municipality",
            }
        )

    counts = Counter(row["level_code"] for row in seed_rows)
    if dict(counts) != EXPECTED_COUNTS:
        raise RuntimeError(f"Unexpected Spain seed counts: {dict(counts)!r} != {EXPECTED_COUNTS!r}")
    return seed_rows


def write_seed(rows: list[dict[str, str]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "level_code",
                "admin_code",
                "display_name",
                "territory_name",
                "territory_wikidata_id",
                "territory_type",
                "parent_level_code",
                "parent_admin_code",
                "source",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    args = parse_args()
    municipalities, provinces = load_ine_rows(args.ine_workbook)
    snapshot_qids, snapshot_names_by_qid = load_snapshot_metadata(args.territory_sql)
    province_matches = load_wikidata_matches(
        [province["admin_code"] for province in provinces],
        chunk_size=args.chunk_size,
        sleep_seconds=args.sleep_seconds,
    )
    municipality_matches = load_wikidata_matches(
        [municipality["admin_code"] for municipality in municipalities],
        chunk_size=args.chunk_size,
        sleep_seconds=args.sleep_seconds,
    )
    seed_rows = build_seed_rows(
        municipalities,
        provinces,
        province_matches,
        municipality_matches,
        snapshot_qids,
        snapshot_names_by_qid,
    )
    write_seed(seed_rows, args.output)
    print(f"Wrote {len(seed_rows)} Spain administrative rows to {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
